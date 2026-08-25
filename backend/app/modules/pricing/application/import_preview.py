from __future__ import annotations

from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.platform.security.authorization import (
    AuthorizationPolicyPort,
    AuthorizationRequest,
    AuthorizationResource,
)
from app.platform.security.capabilities import Capability
from app.platform.security.context import ActorContext, ActorKind, DataClassification

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 50 * 1024 * 1024
MAX_ROWS = 10_000
MAX_ERRORS = 100

_HEADER_ALIASES = {
    "code": {"code", "référence", "reference", "ref", "n°", "no"},
    "designation": {"désignation", "designation", "libellé", "libelle", "description"},
    "unit": {"unité", "unite", "unit", "u"},
    "quantity": {"quantité", "quantite", "quantity", "qty"},
    "unit_price": {"prix unitaire", "pu", "unit price", "unit_price"},
    "total_price": {"montant", "total", "prix total", "total price", "total_price"},
}


@dataclass(frozen=True, slots=True)
class PricingImportRow:
    row_number: int
    code: str | None
    designation: str | None
    unit: str | None
    quantity_decimal: str | None
    unit_price_minor: int | None
    total_minor: int | None
    errors: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class PricingImportPreview:
    case_id: object
    document_kind: str
    filename: str
    row_count: int
    valid_row_count: int
    error_count: int
    total_minor: int
    truncated: bool
    limit_reason: str | None
    rows: tuple[PricingImportRow, ...]


class PricingImportPreviewService:
    def __init__(self, *, policy: AuthorizationPolicyPort) -> None:
        self._policy = policy

    def preview(
        self,
        *,
        actor: ActorContext,
        case_id,
        document_kind: str,
        filename: str,
        content_type: str | None,
        payload: bytes,
    ) -> PricingImportPreview:
        self._authorize(actor=actor, case_id=case_id)
        if len(payload) > MAX_UPLOAD_BYTES:
            raise ValueError("IMPORT_FILE_TOO_LARGE")
        normalized_name = filename.casefold().strip()
        if not normalized_name.endswith(".xlsx") or normalized_name.endswith(".xlsm"):
            raise ValueError("IMPORT_XLSX_REQUIRED")
        accepted_content_types = {
            None,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        if content_type not in accepted_content_types:
            raise ValueError("IMPORT_CONTENT_TYPE_REJECTED")
        try:
            with ZipFile(BytesIO(payload)) as archive:
                if any(name.lower().endswith("vbaproject.bin") for name in archive.namelist()):
                    raise ValueError("IMPORT_MACROS_REJECTED")
                if sum(info.file_size for info in archive.infolist()) > MAX_UNCOMPRESSED_BYTES:
                    raise ValueError("IMPORT_ARCHIVE_TOO_LARGE")
                _check_archive_uncompressed_size(archive)
            workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        except (BadZipFile, OSError, ValueError) as error:
            if isinstance(error, ValueError) and str(error).startswith("IMPORT_"):
                raise
            raise ValueError("IMPORT_WORKBOOK_INVALID") from error
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                raise ValueError("IMPORT_EMPTY_WORKBOOK")
            header_map = _resolve_headers(header)
            if "designation" not in header_map:
                raise ValueError("IMPORT_DESIGNATION_COLUMN_REQUIRED")
            parsed: list[PricingImportRow] = []
            total_minor = 0
            errors = 0
            rows_budget_reached = False
            errors_budget_reached = False
            for row_number, values in enumerate(rows, start=2):
                if row_number > MAX_ROWS + 1:
                    rows_budget_reached = True
                    break
                if not any(value not in (None, "") for value in values):
                    continue
                item = _parse_row(row_number=row_number, values=values, header_map=header_map)
                parsed.append(item)
                errors += len(item.errors)
                if not item.errors and item.total_minor is not None:
                    total_minor += item.total_minor
                if errors >= MAX_ERRORS:
                    errors_budget_reached = True
                    break
            truncated = (rows_budget_reached or errors_budget_reached) and _has_remaining_rows(
                rows
            )
            limit_reason: str | None = None
            if truncated:
                limit_reason = "ROW_LIMIT" if rows_budget_reached else "ERROR_LIMIT"
            return PricingImportPreview(
                case_id=case_id,
                document_kind=document_kind,
                filename=filename,
                row_count=len(parsed),
                valid_row_count=sum(not item.errors for item in parsed),
                error_count=errors,
                total_minor=total_minor,
                truncated=truncated,
                limit_reason=limit_reason,
                rows=tuple(parsed),
            )
        finally:
            workbook.close()

    def _authorize(self, *, actor: ActorContext, case_id) -> None:
        if actor.actor_kind is not ActorKind.PATRON_ADMIN or actor.membership_id is None:
            raise PermissionError("FINANCIAL_REPORT_PATRON_REQUIRED")
        decision = self._policy.authorize(
            context=actor,
            request=AuthorizationRequest(
                action=Capability.FINANCIAL_REPORT_LINE_WRITE,
                resource=AuthorizationResource(
                    resource_type="PRICING_IMPORT",
                    resource_id=case_id,
                    tenant_id=actor.tenant_id,
                    case_id=case_id,
                    classification=DataClassification.FINANCIAL_PRIVATE,
                ),
                evaluated_at=actor.authenticated_at,
            ),
        )
        if not decision.allowed:
            raise PermissionError(decision.code)


def _check_archive_uncompressed_size(archive: ZipFile) -> None:
    total_bytes = 0
    for info in archive.infolist():
        if getattr(info, "is_dir", lambda: False)():
            continue
        with archive.open(info) as member:
            while chunk := member.read(1024 * 1024):
                total_bytes += len(chunk)
                if total_bytes > MAX_UNCOMPRESSED_BYTES:
                    raise ValueError("IMPORT_ARCHIVE_TOO_LARGE")


def _has_remaining_rows(rows) -> bool:
    return any(
        any(value not in (None, "") for value in values) for values in rows
    )


def _normalize_header(value: object) -> str:
    return " ".join(str(value or "").casefold().strip().split())


def _resolve_headers(values: tuple[object, ...]) -> dict[str, int]:
    resolved: dict[str, int] = {}
    aliases = {
        key: {_normalize_header(alias) for alias in names} for key, names in _HEADER_ALIASES.items()
    }
    for index, value in enumerate(values):
        normalized = _normalize_header(value)
        for field, field_aliases in aliases.items():
            if normalized in field_aliases and field not in resolved:
                resolved[field] = index
                break
    return resolved


def _value(values: tuple[object, ...], header_map: dict[str, int], field: str) -> object:
    index = header_map.get(field)
    return values[index] if index is not None and index < len(values) else None


def _parse_row(
    *, row_number: int, values: tuple[object, ...], header_map: dict[str, int]
) -> PricingImportRow:
    code = _text(_value(values, header_map, "code"))
    designation = _text(_value(values, header_map, "designation"))
    unit = _text(_value(values, header_map, "unit"))
    quantity = _decimal_text(_value(values, header_map, "quantity"), default="1")
    unit_price = _minor(_value(values, header_map, "unit_price"))
    total = _minor(_value(values, header_map, "total_price"))
    errors: list[str] = []
    if not designation:
        errors.append("DESIGNATION_REQUIRED")
    if quantity is None:
        errors.append("QUANTITY_INVALID")
    if total is None and unit_price is None:
        errors.append("PRICE_REQUIRED")
    if unit_price is not None and quantity is not None:
        calculated_total = _calculate_total_minor(quantity=quantity, unit_price_minor=unit_price)
        if total is None:
            total = calculated_total
        elif total != calculated_total:
            errors.append("TOTAL_PRICE_MISMATCH")
    return PricingImportRow(
        row_number=row_number,
        code=code,
        designation=designation,
        unit=unit,
        quantity_decimal=quantity,
        unit_price_minor=unit_price,
        total_minor=total,
        errors=tuple(errors),
    )


def _text(value: object) -> str | None:
    if value in (None, ""):
        return None
    return str(value).strip()[:500] or None


def _decimal_text(value: object, *, default: str | None = None) -> str | None:
    if value in (None, ""):
        return default
    decimal = _decimal(value)
    if decimal is None or decimal < 0:
        return None
    return format(decimal, "f")


def _minor(value: object) -> int | None:
    if value in (None, ""):
        return None
    decimal = _decimal(value)
    if decimal is None or decimal < 0:
        return None
    return int((decimal * 100).to_integral_value(rounding=ROUND_HALF_UP))


def _calculate_total_minor(*, quantity: str, unit_price_minor: int) -> int:
    return int(
        (Decimal(quantity) * Decimal(unit_price_minor)).to_integral_value(rounding=ROUND_HALF_UP)
    )


def _decimal(value: object) -> Decimal | None:
    text = str(value).strip().replace("\u00a0", "").replace(" ", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        decimal = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return decimal if decimal.is_finite() else None
